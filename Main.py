import re

import Objects as myObjects
import Regexes as myRegex
import Request as myRequests
from openpyxl import load_workbook
from openpyxl import Workbook


def return_first_match(regex, text):
    x = re.findall(regex, text)
    result = x[0] if len(x) > 0 else ''

    return result


def return_all_matches(regex, text):
    result = re.findall(regex, text)

    return result


def has_number(inputString):
    return any(char.isdigit() for char in inputString)


# Reading.
workbook = load_workbook(filename="UK_1.xlsx")
sheet = workbook.active

# Writing.
writing = Workbook()
sheet_for_write = writing.active

sheet_for_write["A1"] = "Name"
sheet_for_write["B1"] = "Address"
sheet_for_write["C1"] = "County"
sheet_for_write["D1"] = "Postcode"
sheet_for_write["E1"] = "Telephone"
sheet_for_write["F1"] = "Email"
sheet_for_write["G1"] = "Chairman_Email"
sheet_for_write["H1"] = "Secretary_Email"
sheet_for_write["I1"] = "Google_Map"
sheet_for_write["J1"] = "Link"

write_index = 2
for row in sheet.values:
    name = ''
    address = ''
    county = ''
    postal_code = ''
    telephone = ''
    extra_email = ''
    chairman_email = ''
    secretary_email = ''
    google_map = ''
    link = ''

    if row[9] == '1':
        continue
    if row[8] == '':
        continue
    if row[8] == 'Link':
        continue

    print('Loading url (' + str(write_index) + '/' + str(sheet.max_row) + ')... ' + row[8])

    main_request = myObjects.RequestObject()
    main_request.method = 'GET'
    main_request.url = row[8] + 'Aboutus'
    main_source = myRequests.selenium_request(main_request)

    if 'ECB Error Page' in main_source:
        sheet_for_write['A' + str(write_index)] = "Non-existent"
        sheet_for_write['J' + str(write_index)] = row[8]
        write_index += 1
        continue

    list_of = return_all_matches(myRegex.emails, main_source)

    for element in list_of:
        if 'chairman' in element[0].lower():
            chairman_email = element[1]

        if 'secretary' in element[0].lower():
            secretary_email = element[1]

    joining_message = return_first_match(myRegex.joining_message, main_source)
    address_information = return_first_match(myRegex.address_part, main_source)

    parts_of_address = return_all_matches(myRegex.address, address_information)

    if len(parts_of_address) != 0:
        # Remove last element is it's telephone.
        if 'Telephone' in parts_of_address[-1]:
            parts_of_address.pop()

        # Check if last element is postal code.
        if has_number(parts_of_address[-1]):
            postal_code = parts_of_address[-1]
            parts_of_address.pop()

        # Get county.
        if len(parts_of_address) != 0:
            county = parts_of_address[-1]
            parts_of_address.pop()

        # Construct address.
        if len(parts_of_address) != 0:
            address = ' '.join(parts_of_address)

    postal_code_for_maps = postal_code
    if postal_code == '':
        postal_code_for_maps = address

        if address == '':
            postal_code_for_maps = county

    name = return_first_match(myRegex.name, main_source)
    telephone = return_first_match(myRegex.mobile_phone, main_source)
    extra_email = return_first_match(myRegex.extra_email, joining_message)
    google_map = 'http://maps.google.co.uk/?q=' + postal_code_for_maps
    link = row[8]

    sheet_for_write['A' + str(write_index)] = name
    sheet_for_write['B' + str(write_index)] = address
    sheet_for_write['C' + str(write_index)] = county
    sheet_for_write['D' + str(write_index)] = postal_code
    sheet_for_write['E' + str(write_index)] = telephone
    sheet_for_write['F' + str(write_index)] = extra_email
    sheet_for_write['G' + str(write_index)] = chairman_email
    sheet_for_write['H' + str(write_index)] = secretary_email
    sheet_for_write['I' + str(write_index)] = google_map
    sheet_for_write['J' + str(write_index)] = link

    write_index += 1

writing.save(filename='Result.xlsx')
myRequests.close_browser()
